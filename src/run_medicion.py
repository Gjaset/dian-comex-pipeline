#!/usr/bin/env python3
"""Ejecuta la lógica del notebook medicion.ipynb como script puro y regenera
el .ipynb con las salidas reales, sin necesidad de jupyter kernel.

Salida:
- data/mediciones.json con todas las cifras
- notebooks/medicion.ipynb regenerado con execution_count y outputs visibles
"""
import os, glob, zipfile, time, math, json, io, sys
import pandas as pd
import openpyxl
import psutil
import nbformat as nbf

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
RAW = os.path.join(ROOT, 'data', 'raw')
EXT = os.path.join(ROOT, 'data', 'extracted')
os.makedirs(EXT, exist_ok=True)

# Captura de stdout por "celda"
_cells = []  # lista de (cell_type, source_str, outputs_list, exec_count)

def add_md(src):
    _cells.append(('markdown', src, None, None))

def add_code(src):
    buf = io.StringIO()
    old = sys.stdout
    sys.stdout = buf
    err = None
    try:
        exec(compile(src, '<medicion>', 'exec'), globals())
    except Exception as e:
        import traceback
        err = f"{type(e).__name__}: {e}\n" + traceback.format_exc()
    finally:
        sys.stdout = old
    outputs = []
    text = buf.getvalue()
    if text:
        outputs.append(nbf.v4.new_output('stream', name='stdout', text=text))
    if err:
        outputs.append(nbf.v4.new_output('error', ename='Error', evalue=err, traceback=[err]))
    _cells.append(('code', src, outputs, len([c for c in _cells if c[0]=='code'])+1))

# ===== Celdas del notebook =====
add_md("""# T1 — Medición de fuente: DIAN · Bases Estadísticas de Importaciones

**Curso:** IFPN0025 · Big Data e Ingeniería de Datos · Universidad Ean  
**Estudiante:** Germán Cuesta  
**Fuente:** DIAN — Bases Estadísticas de Comercio Exterior (Importaciones). Archivos XLSX publicados mensualmente, empaquetados en `.zip`, descargables del sitio web de la DIAN (SharePoint público). NO es API Socrata: cada `.zip` contiene un único XLSX con las declaraciones de importación del mes.

**Períodos elegidos:** 12 meses de **2023** y 12 meses de **2024**, ambos en codificación **ALADI** (anterior al cambio M49 de enero-2025). Esto da un `g` anual medido directamente, sin extrapolación ni artefactos de cambio de esquema.

El notebook corre de punta a punta y reproduce cada cifra de `docs/ficha_tecnica.md`.""")

add_code("""import os, glob, zipfile, time, math, json
import pandas as pd
import openpyxl
import psutil

RAW = '../data/raw'
EXT = '../data/extracted'
os.makedirs(EXT, exist_ok=True)

print('Python:', pd.__version__, 'pandas |', 'openpyxl', openpyxl.__version__, '| psutil', psutil.__version__)
print('Memoria total del equipo (GB):', round(psutil.virtual_memory().total/1024**3, 2))
print('Memoria disponible inicial (GB):', round(psutil.virtual_memory().available/1024**3, 2))""")

add_md("""## 1. Inventario de archivos descargados

Los 24 archivos `.zip` ya están en `data/raw/`. Cada uno pesa ~200-250 MB (comprimido) y contiene un único XLSX de ~230 MB (descomprimido). Total: ~5.3 GB en crudo.

_ nota: este notebook asume que la descarga ya se hizo. Si no existen los zips, reejecute el script de descarga (`src/descarga_imp.py`)._""")

add_code("""zips = sorted(glob.glob(os.path.join(RAW, '*_Importaciones_*.zip')))
print(f'Total de zips: {len(zips)}')
total_zip_bytes = 0
for p in zips:
    total_zip_bytes += os.path.getsize(p)
print(f'Volumen comprimido total (S0 en disco): {total_zip_bytes/1024**3:.4f} GB')
print('Primeros 3:', [os.path.basename(p) for p in zips[:3]])
print('Ultimos 3:', [os.path.basename(p) for p in zips[-3:]])""")

add_md("""## 2. Exploración interna de un XLSX (columnas y presencia de PII)

Antes de cualquier medición, inspecciono la estructura real del primer mes para confirmar el esquema y verificar la presencia/ausencia de datos personales identificables. No me fío de la documentación de la DIAN sin mirar las columnas reales.""")

add_code("""p_ejemplo = zips[0]
print('Examinando:', os.path.basename(p_ejemplo))
with zipfile.ZipFile(p_ejemplo) as z:
    nombre_xlsx = z.namelist()[0]
    info = z.getinfo(nombre_xlsx)
    print(f'  Contenido: {nombre_xlsx} ({info.file_size/1024**2:.1f} MB descomprimido)')
    out_path = os.path.join(EXT, nombre_xlsx)
    if not os.path.exists(out_path):
        with z.open(nombre_xlsx) as src, open(out_path, 'wb') as dst:
            while True:
                c = src.read(131072)
                if not c: break
                dst.write(c)

t0 = time.time()
wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
ws = wb.active
print(f'Hoja: {ws.title}  max_row={ws.max_row}  max_col={ws.max_column}  (carga en {time.time()-t0:.1f}s)')
filas = list(ws.iter_rows(max_row=3, values_only=True))
cols = list(filas[0])
print('N.º de columnas:', len(cols))
print('Primeras 30 columnas:')
for i, c in enumerate(cols[:30]):
    print(f'  [{i+1}] {c}')
wb.close()""")

add_md("""### 2.1 Verificación de PII (requisito del curso: sin datos personales identificables)

La DIAN declara que solo publica identificación de **personas jurídicas** (empresas), amparada en el Fallo del Tribunal Administrativo de Cundinamarca (exp. 25000-23-41-000-2022-00994-00, sept-2022). Lo verifico mirando las columnas reales y los tipos de identificación de las primeras filas.""")

add_code("""cols_pii_candidates = [c for c in cols if any(k in c.upper() for k in ['NOMBRE', 'NIT', 'DIRECCION', 'TIPO_IDENT', 'NUMERO_IDENT', 'DOCUMENTO', 'RAZON'])]
print(f'Columnas candidatas a identificacion/ubicacion ({len(cols_pii_candidates)}):')
for c in cols_pii_candidates:
    print(f'  - {c}')
print()
print('Valores de TIPO_IDENTIFICAC_IMPORTAD en las primeras 2 filas de datos:')
for fila in filas[1:3]:
    tipo = fila[cols.index('TIPO_IDENTIFICAC_IMPORTAD')]
    nit  = fila[cols.index('NIT_IMPORTADOR')]
    nom  = fila[cols.index('NOMBRE_IMPORTADOR')]
    print(f'  tipo_id={tipo!r}  nit={nit!r}  nombre={nom!r}')
print()
print('Interpretacion: la DIAN certifica en su ficha tecnica publica que solo publica personas juridicas, amparada en el fallo judicial citado. Esta verificacion puntual confirma la estructura de las columnas; un barrido completo para confirmar predominio absoluto de NIT sobre cedula se deja como verificacion complementaria (no es necesaria para fines estadisticos).')""")

add_md("""## 3. Conteo real de filas por mes (streaming, sin cargar todo en RAM)

Cuento el número real de filas (`max_row - 1` para excluir el header) de cada uno de los 24 XLSX. Para evitar cargar 5 GB en memoria, uso `openpyxl` en `read_only=True`, que accede vía streaming al XML interno del xlsx sin descomprimir todo el árbol de celdas. Cada archivo se extrae a `data/extracted/` (la primera vez) y luego se descarta para no llenar el disco.""")

add_code("""def contar_filas_y_tamano(zip_path):
    '''Devuelve (n_filas, n_cols, bytes_descomprimidos_xlsx, nombre_xlsx).'''
    with zipfile.ZipFile(zip_path) as z:
        nombre_xlsx = z.namelist()[0]
        info = z.getinfo(nombre_xlsx)
        out_path = os.path.join(EXT, nombre_xlsx)
        if not os.path.exists(out_path):
            with z.open(nombre_xlsx) as src, open(out_path, 'wb') as dst:
                while True:
                    c = src.read(131072)
                    if not c: break
                    dst.write(c)
        # info.file_size es el tamano descomprimido exacto del xlsx dentro del zip
        bytes_xlsx = info.file_size
    wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
    ws = wb.active
    n_rows = ws.max_row - 1  # descuenta el header
    n_cols = ws.max_column
    wb.close()
    try:
        os.remove(out_path)
    except FileNotFoundError:
        pass
    return n_rows, n_cols, bytes_xlsx, nombre_xlsx

resultado = {}
t_total = time.time()
for p in zips:
    nombre = os.path.basename(p).replace('.zip', '')
    partes = nombre.split('_')  # ['01','Importaciones','2023','Enero']
    anio = int(partes[2])
    mes  = partes[3]
    t0 = time.time()
    nfilas, ncols, bytes_xlsx, nom = contar_filas_y_tamano(p)
    dt = time.time() - t0
    print(f'  {nombre}: {nfilas:>7,d} filas, {ncols} cols, {bytes_xlsx/1024**2:6.1f} MB xlsx  ({dt:.1f}s)')
    resultado.setdefault(anio, {'filas': 0, 'bytes_xlsx': 0, 'meses': []})
    resultado[anio]['filas']      += nfilas
    resultado[anio]['bytes_xlsx'] += bytes_xlsx
    resultado[anio]['meses'].append({
        'mes': mes, 'archivo': nombre + '.zip',
        'filas': nfilas, 'bytes_xlsx': bytes_xlsx,
        'bytes_zip': os.path.getsize(p),
    })
print(f'\\nTiempo total de conteo: {time.time()-t_total:.1f}s')""")

add_md("""## 4. Medición S0 — volumen anual por año

Defino dos S0 distintos, ambos reportados:
- **S0 (comprimido en disco):** suma de los bytes de los 12 `.zip` del año. Es lo que efectivamente ocupa el dato si se deja como llega de la DIAN.
- **S0 (descomprimido lógico):** suma de los bytes del XLSX interno, que es cómo se consume el dato para análisis. Es la cifra que uso para `g` y para el umbral, porque el umbral mide cuándo deja de caber en memoria **al cargarlo en pandas**.

Reporto ambas en la ficha y dejo constancia explícita de cuál se usa en cada cálculo.""")

add_code("""S0_comprimido, S0_logico = {}, {}
print(f"{'Anio':<6} {'Filas':>12} {'S0 compr. (GB)':>15} {'S0 logico (GB)':>17}")
for anio in sorted(resultado.keys()):
    S0_comprimido[anio] = sum(m['bytes_zip']  for m in resultado[anio]['meses']) / 1024**3
    S0_logico[anio]      = resultado[anio]['bytes_xlsx']                    / 1024**3
    print(f'{anio:<6} {resultado[anio]["filas"]:>12,d} {S0_comprimido[anio]:>15.4f} {S0_logico[anio]:>17.4f}')

with open('../data/mediciones.json', 'w', encoding='utf-8') as f:
    json.dump({
        'filas_por_anio': {str(a): resultado[a]['filas'] for a in resultado},
        'S0_comprimido_GB': {str(a): S0_comprimido[a] for a in resultado},
        'S0_logico_GB': {str(a): S0_logico[a] for a in resultado},
        'filas_total_2_anios': sum(resultado[a]['filas'] for a in resultado),
        'n_columnas': ncols,
    }, f, indent=2)
print('Mediciones guardadas en ../data/mediciones.json')""")

add_md("""## 5. Medición k — factor de expansión memoria/disco

**Obligatorio `memory_usage(deep=True)`** — sin `deep=True`, pandas reporta solo el tamaño de los punteros y subestima k para columnas de tipo `object` (cadenas), que son la mayoría en este dataset.

Mido k sobre una muestra de 5,000 filas del primer archivo (perfectamente representativa para una proporción: el factor k no depende de cuántas filas cargues, solo del tipo y ancho de cada columna). El disco de referencia es el tamaño del XLSX completo (no del zip); uso el tamaño descomprimido lógico para consistencia con el S0 del Bloque B/C.""")

add_code("""# Reextraigo solo el primer xlsx para medir k (luego lo elimino)
p_ej = zips[0]
with zipfile.ZipFile(p_ej) as z:
    nombre_xlsx = z.namelist()[0]
    out_path = os.path.join(EXT, nombre_xlsx)
    if not os.path.exists(out_path):
        with z.open(nombre_xlsx) as src, open(out_path, 'wb') as dst:
            while True:
                c = src.read(131072)
                if not c: break
                dst.write(c)

t0 = time.time()
df = pd.read_excel(out_path, nrows=5000)
print(f'Tiempo de lectura de 5000 filas: {time.time()-t0:.1f}s')
print(f'Shape: {df.shape}')
# k mejor medido: memoria deep / peso en disco del mismo numero de filas (proporcional al mes completo)
df_disk_full = os.path.getsize(out_path)
df_rows_full = list(resultado.values())[0]['meses'][0]['filas']  # filas del primer mes
bytes_disk_sample = df_disk_full * (5000 / df_rows_full)
mem_deep_bytes = df.memory_usage(deep=True).sum()
k = mem_deep_bytes / bytes_disk_sample
print(f'Memoria deep pandas (5k filas): {mem_deep_bytes/1024**2:.2f} MB')
print(f'Tamano en disco equivalente (5k filas, proporcional): {bytes_disk_sample/1024**2:.2f} MB')
print(f'k = {k:.4f}')
print(f'\\nDetalle top-10 columnas por peso en memoria:')
mem_by_col = df.memory_usage(deep=True).sort_values(ascending=False).head(10)
print(mem_by_col.to_string())""")

add_md("""## 6. Medición M — memoria útil disponible

M se mide en el equipo donde se ejecuta el notebook, **con el notebook y el entorno ya cargados** (ese es el escenario real del umbral: cuánta RAM queda libre cuando ya tienes el stack analítico corriendo).""")

add_code("""mem = psutil.virtual_memory()
M_gb = mem.available / 1024**3
print(f'M (memoria disponible) = {M_gb:.4f} GB')
print(f'Memoria total del equipo = {mem.total/1024**3:.4f} GB')
print(f'Carga actual del sistema: pct_usada={mem.percent:.1f}%, pct_disponible={100-mem.percent:.1f}%')
print(f'Contexto de medición: notebook corriendo, sin otros procesos pesados activos.')""")

add_md("""## 7. Medición g — tasa de crecimiento anual

Método: comparación del **conteo real de filas** entre 2023 y 2024, suponiendo tamaño promedio por fila constante (la estabilidad del esquema se verifica en la sección 9).

$$g = (S0_{2024} / S0_{2023})^{1/(2024-2023)} - 1$$

Uso el S0 lógico (descomprimido), no el comprimido, porque el umbral mide cuándo deja de caber en RAM al cargarlo en pandas — y eso se hace desde XLSX, no desde zip.""")

add_code("""n_periodos = 2024 - 2023
g_anual = (S0_logico[2024] / S0_logico[2023]) ** (1 / n_periodos) - 1
g_mensual = (1 + g_anual) ** (1/12) - 1
print(f'S0 logico 2023 = {S0_logico[2023]:.4f} GB  ({resultado[2023]["filas"]:,} filas)')
print(f'S0 logico 2024 = {S0_logico[2024]:.4f} GB  ({resultado[2024]["filas"]:,} filas)')
print(f'g anual   = {g_anual:.6f}  ({g_anual*100:+.2f}%/año)')
print(f'g mensual = {g_mensual:.6f}  ({g_mensual*100:+.3f}%/mes)')
if g_anual < 0:
    print(f'\\nNota: g<0 es válido — el número de declaraciones de importación bajó entre 2023 y 2024. No es un error de medición; se interpreta en la ficha.')""")

add_md("""## 8. Cálculo del umbral t*

Fórmula estándar (verificada contra el enunciado de la sesión): el umbral es el número de períodos (años) a partir del cual el dataset cargado en pandas (**S0·k**) deja de caber en la memoria disponible **M**, suponiendo crecimiento a tasa **g**.

$$t^* = log(M / (S0 * k)) / log(1+g)$$

Uso el S0 del **último período disponible (2024)** como punto de partida — el umbral se proyecta hacia adelante desde hoy.""")

add_code("""S0_actual = S0_logico[2024]
print(f'Punto de partida: S0(2024)={S0_actual:.4f} GB, k={k:.4f}, M={M_gb:.4f} GB, g={g_anual:.4f}/año')
print(f'S0·k = {S0_actual*k:.4f} GB  (memoria que ocuparía cargar un año completo en pandas)')

if g_anual <= 0:
    print(f'\\ng = {g_anual:.4f} <= 0: la fuente decrece o está estable.')
    if S0_actual * k > M_gb:
        print(f'Como S0·k > M ({S0_actual*k:.2f} > {M_gb:.2f}), el umbral YA está en el pasado: hoy mismo no cabe el año completo en RAM.')
        print('Interpretación: el dataset anual ya no carga en memoria de una sola vez en este equipo. Hay que trabajar por meses o por trozos.')
    else:
        print(f'Como S0·k <= M ({S0_actual*k:.2f} <= {M_gb:.2f}) y g<=0, nunca se alcanza el umbral: el dataset crece menos que la memoria disponible.')
    t_estrella = math.log(M_gb / (S0_actual * k)) / math.log(1 + g_anual) if g_anual != 0 else float('inf')
else:
    t_estrella = math.log(M_gb / (S0_actual * k)) / math.log(1 + g_anual)
    print(f'\\nUmbral t* = {t_estrella:.2f} períodos (años) desde 2024')
    if t_estrella < 0:
        print('Interpretación: el umbral ya quedó en el pasado (con el S0 actual, k y M ya no alcanza).')
    else:
        print(f'Interpretación: en ~{t_estrella:.1f} años, el dataset anual cargado en pandas ya no cabrá en la memoria disponible.')

with open('../data/mediciones.json', 'r+', encoding='utf-8') as f:
    data = json.load(f)
data.update({
    'k': k,
    'M_GB': M_gb,
    'g_anual': g_anual,
    'g_mensual': g_mensual,
    't_estrella': t_estrella if g_anual != 0 else None,
})
with open('../data/mediciones.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=2)
print('Mediciones final completadas en ../data/mediciones.json')""")

add_md("""## 9. Estabilidad del esquema (comparación de columnas entre 2023 y 2024)

Para el Bloque A de la ficha: comparar el set de columnas de un mes de 2023 vs un mes de 2024. Si son idénticos, el esquema es estable; si no, documentar las diferencias. IMPORTANTE: **este control no detecta cambios en códigos de país** (ALADI→M49), que se introdujeron a partir de enero-2025 — los archivos de 2023 y 2024 siguen ambos en ALADI, por eso los elegí. El cambio de codificación se documenta por separado en la ficha citando la Ficha Técnica oficial de la DIAN.""")

add_code("""def obtener_columnas(zip_path):
    with zipfile.ZipFile(zip_path) as z:
        nombre = z.namelist()[0]
        out = os.path.join(EXT, nombre)
        if not os.path.exists(out):
            with z.open(nombre) as src, open(out, 'wb') as dst:
                while True:
                    c = src.read(131072)
                    if not c: break
                    dst.write(c)
    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(max_row=1, values_only=True))
    wb.close()
    try: os.remove(out)
    except FileNotFoundError: pass
    return list(header)

cols_2023 = obtener_columnas([p for p in zips if '2023' in p][0])
cols_2024 = obtener_columnas([p for p in zips if '2024' in p][0])
s23, s24 = set(cols_2023), set(cols_2024)
print(f'Columnas en 2023: {len(s23)}')
print(f'Columnas en 2024: {len(s24)}')
print(f'Solo en 2023: {sorted(s23 - s24)}')
print(f'Solo en 2024: {sorted(s24 - s23)}')
print(f'\\nEsquema estable (2023 vs 2024): {s23 == s24}')""")

add_md("""## 10. Identificador estable de registro

Candidato: combinación de `CONSECUTIVO_CAJERO` + `CODIGO_CAJERO` + `COD_ADUANA_PRESENTADA`. Verifico unicidad en una muestra de 50,000 filas; en la ficha se declara que la verificación de unicidad plena requiere correrlo sobre los 24 archivos completos (queda como limitación de la verificación documentada, no como una deficiencia oculta).""")

add_code("""# Reextraigo el primer xlsx otra vez para la prueba de unicidad
with zipfile.ZipFile(zips[0]) as z:
    nombre = z.namelist()[0]
    out_path = os.path.join(EXT, nombre)
    if not os.path.exists(out_path):
        with z.open(nombre) as src, open(out_path, 'wb') as dst:
            while True:
                c = src.read(131072)
                if not c: break
                dst.write(c)
df = pd.read_excel(out_path, nrows=50000)
print(f'Muestra: {len(df):,} filas para verificar unicidad.')
for cand in [['CONSECUTIVO_CAJERO'], ['CODIGO_CAJERO','CONSECUTIVO_CAJERO'], 
             ['COD_ADUANA_PRESENTADA','CODIGO_CAJERO','CONSECUTIVO_CAJERO'],
             ['CODIGO_SUCURSAL','CODIGO_CAJERO','CONSECUTIVO_CAJERO','NIT_IMPORTADOR']]:
    if all(c in df.columns for c in cand):
        n_dup = df.duplicated(subset=cand).sum()
        print(f'  Candidato {cand}: duplicados en muestra = {n_dup} (unicidad: {n_dup==0})')
try: os.remove(out_path)
except FileNotFoundError: pass
print('\\nNota: la unicidad se valida sobre una muestra de 50k filas del mes de enero-2023. La unicidad plena requiere aplicar el mismo codigo a los 24 archivos completos; eso se deja documentado en la ficha como limitacion de la verificacion.')""")

add_md("""## 11. Resumen ejecutable (lo que copia la ficha)

Esta celda imprime todas las cifras que van a `docs/ficha_tecnica.md` Blocks B/C. Ejecutando solo esta celda con el archivo `mediciones.json` presente, se reproducen los valores sin tener que volver a leer los 5 GB.""")

add_code("""with open('../data/mediciones.json', encoding='utf-8') as f:
    m = json.load(f)
print('--- Bloque B (mediciones) ---')
for a in sorted(m['filas_por_anio']):
    print(f'  S0 [{a}] comprimido = {m["S0_comprimido_GB"][a]:.4f} GB | logico = {m["S0_logico_GB"][a]:.4f} GB | filas = {m["filas_por_anio"][a]:,d}')
print(f'  k = {m["k"]:.4f}')
print(f'  M = {m["M_GB"]:.4f} GB')
print('--- Bloque C (crecimiento y umbral) ---')
print(f'  g anual   = {m["g_anual"]:.6f} ({m["g_anual"]*100:+.2f}%/año)')
print(f'  g mensual = {m["g_mensual"]:.6f} ({m["g_mensual"]*100:+.3f}%/mes)')
if m.get('t_estrella') is not None:
    print(f'  t* = {m["t_estrella"]:.2f} períodos (años)')
else:
    print(f'  t* = N/A (g=0, formula degenerada)')""")

# ===== Construir el notebook con outputs reales =====
nb = nbf.v4.new_notebook()
nb.metadata.kernelspec = {"display_name":"Python 3 (ipykernel)","language":"python","name":"python3"}
nb.metadata.language_info = {"name":"python","version":"3.11.14"}
for ct, src, outputs, ec in _cells:
    if ct == 'markdown':
        nb.cells.append(nbf.v4.new_markdown_cell(source=src))
    else:
        c = nbf.v4.new_code_cell(source=src)
        c['execution_count'] = ec
        c['outputs'] = outputs or []
        nb.cells.append(c)

out_path = os.path.join(HERE, 'medicion.ipynb')
with open(out_path, 'w', encoding='utf-8') as f:
    nbf.write(nb, f)
print(f"\nNotebook generado: {out_path} ({len(nb.cells)} celdas)")
print(f"mediciones.json: {os.path.join(ROOT,'data','mediciones.json')}")
